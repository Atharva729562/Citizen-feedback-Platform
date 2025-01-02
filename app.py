from flask import Flask, request, jsonify
from flask_cors import CORS
from openpyxl import Workbook, load_workbook
import os

app = Flask(__name__)
CORS(app)

# Directory to save Excel files
DATA_DIR = "data"
os.makedirs(DATA_DIR, exist_ok=True)

# File paths for Excel sheets
SURVEY_FILE = os.path.join(DATA_DIR, 'survey_data.xlsx')
FEEDBACK_FILE = os.path.join(DATA_DIR, 'feedback_data.xlsx')

# Initialize Excel files if they don't exist
def initialize_excel(file_path, headers):
    if not os.path.exists(file_path):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(headers)
        workbook.save(file_path)

# Initialize survey and feedback files
initialize_excel(SURVEY_FILE, ["Question1", "Healthcare", "Education", "Public Services", "Priority", "Rating"])
initialize_excel(FEEDBACK_FILE, ["Name", "Contact", "Email", "Comment"])

# Save Survey Data
@app.route('/save-survey', methods=['POST'])
def save_survey():
    survey_data = request.json
    try:
        workbook = load_workbook(SURVEY_FILE)
        sheet = workbook.active
        sheet.append([
            survey_data.get("question1"),
            survey_data.get("healthcare"),
            survey_data.get("education"),
            survey_data.get("publicServices"),
            survey_data.get("priority"),
            survey_data.get("rating")
        ])
        workbook.save(SURVEY_FILE)
        return jsonify({"message": "Survey data saved successfully"}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# Save Feedback Data
@app.route('/save-feedback', methods=['POST'])
def save_feedback():
    feedback_data = request.json
    try:
        workbook = load_workbook(FEEDBACK_FILE)
        sheet = workbook.active
        sheet.append([
            feedback_data.get("name"),
            feedback_data.get("contact"),
            feedback_data.get("email"),
            feedback_data.get("comment")
        ])
        workbook.save(FEEDBACK_FILE)
        return jsonify({"message": "Feedback data saved successfully"}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

if __name__ == '__main__':
    app.run(debug=True, port=5000)
